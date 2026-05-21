"""
Step 6: 明细Excel
输入: 本周4周数据 + 阈值Excel/pkl
输出: 预警明细.xlsx + 贡献pp明细.xlsx
"""
import os
import sys
import pandas as pd
import numpy as np
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(__file__))
from config import (
    CORE_DIMS, CORE_DIM_NAMES, CROSS_LEVELS, EXTRA_DIMS,
    ALERT_METRICS, CONTRIB_METRICS, ALL_METRICS,
    FONT_NAME, COLOR_RED, COLOR_GREEN,
    TARGET_YEAR, TARGET_WEEK,
    load_data, aggregate_weekly, compute_metrics,
    get_week_label, get_output_dir, load_pkl, check_threshold,
    THRESHOLD_FILE, fmt_dau_wan,
)


# ═══════════════════════════════════════════
# 样式
# ═══════════════════════════════════════════
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_FONT = Font(name=FONT_NAME, bold=True, size=9, color='FFFFFF')
CELL_FONT = Font(name=FONT_NAME, size=9)
RED_FONT = Font(name=FONT_NAME, size=9, color=COLOR_RED)
GREEN_FONT = Font(name=FONT_NAME, size=9, color=COLOR_GREEN)
SUMMARY_FONT = Font(name=FONT_NAME, size=9, bold=True)
SUMMARY_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')


def build_analysis_df(weekly_df, dim_fields, total_weekly, thresholds, level_name, metrics_list):
    """构建单个Sheet的分析数据（当前周指标 + 阈值判定）"""
    # 当前周指标
    cur_mask = (weekly_df['iso_year'] == TARGET_YEAR) & (weekly_df['iso_week'] == TARGET_WEEK)

    if dim_fields:
        cur_weekly = weekly_df[cur_mask].groupby(dim_fields, as_index=False)['dau'].sum()
    else:
        cur_weekly = weekly_df[cur_mask].groupby(['iso_year', 'iso_week'], as_index=False)['dau'].sum()

    # 计算指标
    metrics_df = compute_metrics(weekly_df, dim_fields, total_weekly)
    cur_metrics = metrics_df[
        (metrics_df['iso_year'] == TARGET_YEAR) & (metrics_df['iso_week'] == TARGET_WEEK)
    ].copy()

    if cur_metrics.empty:
        return pd.DataFrame()

    # 获取阈值
    level_thresholds = thresholds.get(level_name, {})

    # 构建结果行
    rows = []
    for _, row in cur_metrics.iterrows():
        r = {}
        # 维度列
        for f in dim_fields:
            r[f] = row.get(f, '')
        # DAU(万)
        r['DAU(万)'] = round(row['dau'] / 10000, 0) if row.get('dau') else 0

        # 获取该组合的阈值
        if dim_fields:
            group_key = tuple(row[f] for f in dim_fields)
        else:
            group_key = None

        t = level_thresholds.get(group_key, {}) if isinstance(level_thresholds, dict) and group_key else level_thresholds

        # 指标值 + 阈值判定
        for metric in metrics_list:
            val = row.get(metric)
            r[metric] = round(val, 2) if val is not None else None

            metric_t = t.get(metric, {})
            lo = metric_t.get('lo')
            hi = metric_t.get('hi')
            status, _ = check_threshold(val, lo, hi)
            r[f'{metric}_判定'] = status

        rows.append(r)

    result = pd.DataFrame(rows)

    # 排序：按维度业务排序
    for f in reversed(dim_fields):
        dim_name = None
        for dn, di in {**CORE_DIMS, **EXTRA_DIMS}.items():
            if di['field'] == f:
                dim_name = dn
                break
        if dim_name:
            dim_info = CORE_DIMS.get(dim_name) or EXTRA_DIMS.get(dim_name)
            if dim_info and dim_info.get('values'):
                order = {v: i for i, v in enumerate(dim_info['values'])}
                result['_sort'] = result[f].map(lambda x: order.get(x, 999))
                result = result.sort_values('_sort').drop(columns='_sort')

    return result.reset_index(drop=True)


def generate_alert_summary(df, metrics_list):
    """生成预警明细顶部归因总结"""
    n = len(df)
    if n == 0:
        return ['无数据', '']

    # R1: 触发率概览
    parts = []
    for m in metrics_list:
        col = f'{m}_判定'
        if col in df.columns:
            triggered = df[col].isin(['触发上界', '触发下界']).sum()
            pct = triggered / n * 100
            parts.append(f"{m}触发{triggered}个({pct:.0f}%)")
    r1 = f"触发率概览 — 本Sheet {n}个组合中，{'、'.join(parts)}"

    # R2: 多指标同时触发
    df_temp = df.copy()
    df_temp['_trigger_count'] = 0
    triggered_metrics_list = []
    for m in metrics_list:
        col = f'{m}_判定'
        if col in df_temp.columns:
            df_temp['_trigger_count'] += df_temp[col].isin(['触发上界', '触发下界']).astype(int)
            triggered_metrics_list.append(m)

    multi = df_temp[df_temp['_trigger_count'] >= 2].sort_values('_trigger_count', ascending=False)
    if len(multi) > 0:
        items = []
        for _, row in multi.head(5).iterrows():
            dim_vals = []
            for f in [c for c in df.columns if c not in ['DAU(万)', '_trigger_count'] and not c.endswith('_判定') and c not in metrics_list]:
                dim_vals.append(str(row[f]))
            triggered = []
            for m in metrics_list:
                col = f'{m}_判定'
                if col in row and row[col] in ('触发上界', '触发下界'):
                    triggered.append(m)
            label = ' | '.join(dim_vals)
            items.append(f"{label}({len(triggered)}个指标)")
        r2 = f"2+指标触发的组合：{'、'.join(items)}"
    else:
        r2 = "无2+指标同时触发的组合"

    return [r1, r2]


def generate_contrib_summary(df, dim_fields):
    """生成贡献pp明细顶部归因总结"""
    if df.empty:
        return ['无数据', '']

    lines = []
    for metric in ['WoW贡献pp', 'YoY贡献pp']:
        if metric not in df.columns:
            continue
        sorted_df = df.dropna(subset=[metric]).sort_values(metric)

        drag = sorted_df.head(3)
        pull = sorted_df.tail(3).iloc[::-1]

        drag_parts = []
        for _, row in drag.iterrows():
            label = ' | '.join(str(row[f]) for f in dim_fields)
            drag_parts.append(f"{label} {row[metric]:+.1f}pp")

        pull_parts = []
        for _, row in pull.iterrows():
            label = ' | '.join(str(row[f]) for f in dim_fields)
            pull_parts.append(f"{label} {row[metric]:+.1f}pp")

        line = f"{metric} — 拖累Top3：{'／'.join(drag_parts)}；拉动Top3：{'／'.join(pull_parts)}"
        lines.append(line)

    return lines if lines else ['无数据', '']


def write_detail_excel(output_path, weekly_all, total_weekly, thresholds,
                       metrics_list, summary_func, title_prefix):
    """写明细Excel"""
    print(f"\n生成: {output_path}")

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        for level_name, dim_names in CROSS_LEVELS:
            dim_fields = [CORE_DIMS[d]['field'] for d in dim_names]

            # 聚合到该层级
            weekly_level = weekly_all.groupby(
                ['iso_year', 'iso_week'] + dim_fields, as_index=False
            )['dau'].sum()

            # 构建分析数据
            df = build_analysis_df(weekly_level, dim_fields, total_weekly,
                                   thresholds, level_name, metrics_list)
            if df.empty:
                continue

            sheet_name = level_name[:31]

            # 生成顶部总结
            if summary_func == 'alert':
                summary_lines = generate_alert_summary(df, metrics_list)
            else:
                summary_lines = generate_contrib_summary(df, dim_fields)

            # 写入 — 先空2行给总结
            start_row = len(summary_lines) + 1
            df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row)

            ws = writer.sheets[sheet_name]

            # 写总结行
            for i, line in enumerate(summary_lines):
                cell = ws.cell(row=i + 1, column=1, value=line)
                cell.font = SUMMARY_FONT
                cell.fill = SUMMARY_FILL
                ws.merge_cells(start_row=i + 1, start_column=1,
                               end_row=i + 1, end_column=len(df.columns))

            # 样式
            _style_detail_sheet(ws, df, start_row, dim_fields, metrics_list)

    print(f"  完成: {output_path}")


def _style_detail_sheet(ws, df, header_row, dim_fields, metrics_list):
    """明细Sheet样式"""
    n_rows = len(df)
    n_cols = len(df.columns)
    data_start = header_row + 1

    # 表头样式
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=data_start, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # 数据行样式
    for row_idx in range(data_start + 1, data_start + n_rows + 1):
        for col_idx in range(1, n_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center')

            col_name = df.columns[col_idx - 1]
            val = cell.value

            # 判定列着色
            if col_name.endswith('_判定'):
                if val in ('触发上界', '触发下界'):
                    cell.font = Font(name=FONT_NAME, size=9, color=COLOR_RED, bold=True)
                else:
                    cell.font = CELL_FONT
            # 数值列红正绿负
            elif col_name in metrics_list and isinstance(val, (int, float)):
                if val > 0:
                    cell.font = RED_FONT
                elif val < 0:
                    cell.font = GREEN_FONT
                else:
                    cell.font = CELL_FONT
            else:
                cell.font = CELL_FONT

    # 列宽
    for col_idx in range(1, n_cols + 1):
        col_name = df.columns[col_idx - 1]
        if col_name in dim_fields:
            width = 14
        elif col_name == 'DAU(万)':
            width = 10
        elif col_name.endswith('_判定'):
            width = 10
        else:
            width = 14
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 维度列合并单元格（同值合并）
    for dim_idx, dim_field in enumerate(dim_fields):
        col_idx = dim_idx + 1
        start = data_start + 1
        for row_idx in range(data_start + 2, data_start + n_rows + 2):
            cur_val = ws.cell(row=row_idx, column=col_idx).value
            prev_val = ws.cell(row=row_idx - 1, column=col_idx).value
            if cur_val != prev_val or row_idx == data_start + n_rows + 1:
                end = row_idx - 1 if cur_val != prev_val else row_idx
                if end > start:
                    ws.merge_cells(start_row=start, start_column=col_idx,
                                   end_row=end, end_column=col_idx)
                    ws.cell(row=start, column=col_idx).alignment = Alignment(
                        horizontal='center', vertical='center'
                    )
                start = row_idx


def run(core_data_path):
    """
    core_data_path: 核心3维当周数据文件
    """
    print("=" * 60)
    print("Step 6: 明细Excel")
    print("=" * 60)

    # ── 加载数据 ──
    df = load_data(core_data_path)
    from config import CORE_DIM_FIELDS
    weekly_all = aggregate_weekly(df, CORE_DIM_FIELDS)
    total_weekly = weekly_all.groupby(['iso_year', 'iso_week'], as_index=False)['dau'].sum()
    total_weekly.rename(columns={'dau': 'total_dau'}, inplace=True)

    # ── 加载阈值 ──
    thresholds = load_pkl(THRESHOLD_FILE)

    # ── 输出 ──
    output_dir = get_output_dir()
    week_label = get_week_label(TARGET_YEAR, TARGET_WEEK)

    # 预警明细
    alert_path = os.path.join(output_dir, f'{week_label} TV端全维度叉乘预警明细.xlsx')
    write_detail_excel(alert_path, weekly_all, total_weekly, thresholds,
                       ALERT_METRICS, 'alert', '预警')

    # 贡献pp明细
    contrib_path = os.path.join(output_dir, f'{week_label} TV端贡献pp明细.xlsx')
    write_detail_excel(contrib_path, weekly_all, total_weekly, thresholds,
                       CONTRIB_METRICS, 'contrib', '贡献pp')

    print(f"\n完成！输出目录: {output_dir}")


if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description='TV端DAU明细Excel')
    parser.add_argument('core_data', help='核心3维数据文件 (CSV/Excel)')
    args = parser.parse_args()
    run(args.core_data)
