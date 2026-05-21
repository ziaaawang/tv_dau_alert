"""
Step 2: 阈值Excel + 验证图
输入: thresholds.pkl
输出: DAU异常阈值.xlsx + 验证图 (png)
"""
import os
import sys
import pandas as pd
import numpy as np
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(__file__))
from config import (
    CORE_DIMS, CORE_DIM_NAMES, EXTRA_DIMS, CROSS_LEVELS,
    ALL_METRICS, ALERT_METRICS, CONTRIB_METRICS,
    MANUAL_OVERRIDES, THRESHOLD_FILE, THRESHOLD_EXCEL,
    FONT_NAME, COLOR_RED, COLOR_GREEN,
    load_pkl, save_pkl, get_output_dir,
)


def apply_manual_overrides(thresholds):
    """应用人工微调到整体阈值"""
    if not MANUAL_OVERRIDES:
        return thresholds
    if '整体' not in thresholds:
        return thresholds
    for metric, overrides in MANUAL_OVERRIDES.items():
        if metric in thresholds['整体']:
            if 'lo' in overrides:
                thresholds['整体'][metric]['lo'] = overrides['lo']
            if 'hi' in overrides:
                thresholds['整体'][metric]['hi'] = overrides['hi']
            print(f"  人工微调 {metric}: [{overrides.get('lo', '不变')}, {overrides.get('hi', '不变')}]")
    return thresholds


def write_threshold_excel(thresholds, output_path):
    """将阈值写入Excel"""
    print(f"\n生成阈值Excel: {output_path}")

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    header_font = Font(name=FONT_NAME, bold=True, size=9)
    cell_font = Font(name=FONT_NAME, size=9)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font_white = Font(name=FONT_NAME, bold=True, size=9, color='FFFFFF')

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Sheet 1: 整体阈值
        _write_overall_sheet(writer, thresholds.get('整体', {}),
                            header_font_white, cell_font, header_fill, thin_border)

        # 核心维度 Sheets
        for level_name, dim_names in CROSS_LEVELS:
            level_data = thresholds.get(level_name, {})
            if not level_data:
                continue
            _write_level_sheet(writer, level_name, dim_names, level_data,
                              header_font_white, cell_font, header_fill, thin_border)

        # 补充维度 Sheets
        for dim_name in EXTRA_DIMS:
            key = f"补充_{dim_name}"
            level_data = thresholds.get(key, {})
            if not level_data:
                continue
            _write_level_sheet(writer, dim_name, [dim_name], level_data,
                              header_font_white, cell_font, header_fill, thin_border)

    print(f"  已写入: {output_path}")


def _write_overall_sheet(writer, data, header_font, cell_font, header_fill, border):
    """写整体阈值Sheet"""
    rows = []
    for metric in ALL_METRICS:
        info = data.get(metric, {})
        lo = info.get('lo')
        hi = info.get('hi')
        tr = info.get('trigger_rate')
        rows.append({
            '指标': metric,
            '下界': round(lo, 2) if lo is not None else '-',
            '上界': round(hi, 2) if hi is not None else '-',
            '触发率': f"{tr*100:.0f}%" if tr is not None else '-',
        })
    df = pd.DataFrame(rows)
    df.to_excel(writer, sheet_name='整体阈值', index=False, startrow=0)

    ws = writer.sheets['整体阈值']
    _style_sheet(ws, len(df), 4, header_font, cell_font, header_fill, border)


def _write_level_sheet(writer, sheet_name, dim_names, level_data, header_font, cell_font, header_fill, border):
    """写维度层级阈值Sheet"""
    sheet_name = sheet_name[:31]  # Excel sheet name limit
    rows = []
    for group_keys, metrics_dict in sorted(level_data.items()):
        if not isinstance(group_keys, tuple):
            group_keys = (group_keys,)
        row = {}
        for i, dim_name in enumerate(dim_names):
            row[dim_name] = group_keys[i] if i < len(group_keys) else ''
        for metric in ALL_METRICS:
            info = metrics_dict.get(metric, {})
            lo = info.get('lo')
            hi = info.get('hi')
            tr = info.get('trigger_rate')
            row[f"{metric}_下界"] = round(lo, 2) if lo is not None else '-'
            row[f"{metric}_上界"] = round(hi, 2) if hi is not None else '-'
            row[f"{metric}_触发率"] = f"{tr*100:.0f}%" if tr is not None else '-'
        rows.append(row)

    df = pd.DataFrame(rows)
    df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)

    ws = writer.sheets[sheet_name]
    _style_sheet(ws, len(df), len(df.columns), header_font, cell_font, header_fill, border)


def _style_sheet(ws, n_rows, n_cols, header_font, cell_font, header_fill, border):
    """统一Sheet样式"""
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    for row_idx in range(2, n_rows + 2):
        for col_idx in range(1, n_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = cell_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

    for col_idx in range(1, n_cols + 1):
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ''))
            for r in range(1, min(n_rows + 2, 50))
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 8), 20)


def generate_validation_plots(thresholds, metrics_df, dim_fields_map, output_dir):
    """生成阈值验证图"""
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    matplotlib.rcParams['font.sans-serif'] = ['Microsoft YaHei', 'SimHei']
    matplotlib.rcParams['axes.unicode_minus'] = False

    print(f"\n生成验证图...")

    # 只为单维度画图
    for dim_name in list(CORE_DIM_NAMES) + list(EXTRA_DIMS.keys()):
        # 确定阈值 key
        if dim_name in CORE_DIM_NAMES:
            threshold_key = dim_name
            field = CORE_DIMS[dim_name]['field']
        else:
            threshold_key = f"补充_{dim_name}"
            field = EXTRA_DIMS[dim_name]['field']

        level_thresholds = thresholds.get(threshold_key, {})
        if not level_thresholds:
            continue

        if field not in metrics_df.columns:
            continue

        values = sorted(level_thresholds.keys())
        n_values = len(values)
        n_metrics = len(ALL_METRICS)

        fig, axes = plt.subplots(n_values, n_metrics, figsize=(n_metrics * 3, n_values * 2.5),
                                 squeeze=False)
        fig.suptitle(f'阈值验证 — {dim_name}', fontsize=14, fontweight='bold')

        for i, group_key in enumerate(values):
            label = group_key[0] if isinstance(group_key, tuple) else group_key
            t = level_thresholds[group_key]
            mask = metrics_df[field] == (group_key[0] if isinstance(group_key, tuple) else group_key)
            sub = metrics_df[mask]

            for j, metric in enumerate(ALL_METRICS):
                ax = axes[i][j]
                info = t.get(metric, {})
                lo, hi = info.get('lo'), info.get('hi')
                series = sub[metric].dropna()

                if series.empty:
                    ax.text(0.5, 0.5, '无数据', ha='center', va='center', transform=ax.transAxes)
                    ax.set_title(metric if i == 0 else '', fontsize=8)
                    if j == 0:
                        ax.set_ylabel(label, fontsize=8)
                    continue

                x = range(len(series))
                colors = []
                for v in series:
                    if lo is not None and hi is not None and (v < lo or v > hi):
                        colors.append('#C00000')
                    else:
                        colors.append('#4472C4')

                ax.scatter(x, series.values, c=colors, s=15, zorder=3)

                if lo is not None and hi is not None:
                    ax.axhspan(lo, hi, color='#92D050', alpha=0.2, zorder=1)
                    ax.axhline(lo, color='#92D050', linewidth=0.8, linestyle='--')
                    ax.axhline(hi, color='#92D050', linewidth=0.8, linestyle='--')

                if i == 0:
                    ax.set_title(metric, fontsize=8)
                if j == 0:
                    ax.set_ylabel(label, fontsize=8)
                ax.tick_params(labelsize=6)
                ax.set_xticks([])

        plt.tight_layout(rect=[0, 0, 1, 0.96])
        out_path = os.path.join(output_dir, f'阈值验证_{dim_name}.png')
        plt.savefig(out_path, dpi=150, bbox_inches='tight')
        plt.close(fig)
        print(f"  已生成: {out_path}")


def run(core_data_path=None):
    """
    core_data_path: 可选，原始数据文件路径。提供时会生成验证图。
    """
    print("=" * 60)
    print("Step 2: 阈值Excel + 验证图")
    print("=" * 60)

    # ── 加载阈值 ──
    print(f"\n加载阈值: {THRESHOLD_FILE}")
    thresholds = load_pkl(THRESHOLD_FILE)

    # ── 应用人工微调 ──
    thresholds = apply_manual_overrides(thresholds)

    # ── 写入Excel ──
    output_dir = get_output_dir()
    excel_path = os.path.join(output_dir, 'TV端DAU异常阈值.xlsx')
    write_threshold_excel(thresholds, excel_path)

    # 同时保存一份到 DATA_PATH 供后续脚本读取
    write_threshold_excel(thresholds, THRESHOLD_EXCEL)

    # ── 更新 pkl（含人工微调） ──
    save_pkl(thresholds, THRESHOLD_FILE)

    # ── 生成验证图 ──
    if core_data_path:
        from config import (
            load_data, aggregate_weekly, compute_metrics,
            CORE_DIM_FIELDS, CORE_DIM_NAMES,
        )
        print(f"\n加载原始数据用于生成验证图: {core_data_path}")
        df = load_data(core_data_path)
        weekly_all = aggregate_weekly(df, CORE_DIM_FIELDS)
        total_weekly = weekly_all.groupby(['iso_year', 'iso_week'], as_index=False)['dau'].sum()
        total_weekly.rename(columns={'dau': 'total_dau'}, inplace=True)

        # 为每个单维度计算指标并画图
        for dim_name in CORE_DIM_NAMES:
            dim_field = CORE_DIMS[dim_name]['field']
            weekly_dim = weekly_all.groupby(
                ['iso_year', 'iso_week', dim_field], as_index=False
            )['dau'].sum()
            metrics_dim = compute_metrics(weekly_dim, [dim_field], total_weekly)
            if not metrics_dim.empty:
                generate_validation_plots(thresholds, metrics_dim, {}, output_dir)
                break  # generate_validation_plots 内部遍历所有维度，调一次即可

        # 用完整维度交叉指标调一次，覆盖所有核心维度
        metrics_full = compute_metrics(weekly_all, CORE_DIM_FIELDS, total_weekly)
        if not metrics_full.empty:
            generate_validation_plots(thresholds, metrics_full, {}, output_dir)
    else:
        print("\n未提供原始数据文件，跳过验证图。如需生成，请运行：")
        print("  python step2_threshold_excel.py 数据文件.csv")

    print(f"\n完成！输出目录: {output_dir}")
    return thresholds


if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description='TV端DAU阈值Excel+验证图')
    parser.add_argument('core_data', nargs='?', default=None,
                        help='核心3维数据文件 (CSV/Excel)，提供时生成验证图')
    args = parser.parse_args()
    run(args.core_data)
